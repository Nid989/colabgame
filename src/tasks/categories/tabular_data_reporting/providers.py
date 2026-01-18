"""
Consolidated providers for tabular data reporting category.
Handles file placement, directory creation, setup configuration, and evaluation logic.
"""

import os
from typing import Dict, List, Any, Optional

from ..base import (
    FileProviderInterface,
    ConfigProviderInterface,
    EvaluationProviderInterface,
)
from .evaluators import TabularDataReportingEvaluators


class TabularDataReportingFileProvider(FileProviderInterface):
    """File provider implementation for tabular data reporting tasks."""

    def __init__(self):
        """Initialize tabular data reporting file provider."""
        self.supported_tasks = {
            "simple_data_transfer",
            "basic_data_aggregation",
            "simple_calculation_output",
        }
        self.file_placement_mapping = {
            "simple_data_transfer": "/home/user/Desktop/{filename}",
            "basic_data_aggregation": "/home/user/Desktop/{filename}",
            "simple_calculation_output": "/home/user/Desktop/{filename}",
        }
        self.directory_creation_mapping = {
            "simple_data_transfer": [],
            "basic_data_aggregation": [],
            "simple_calculation_output": [],
        }

    def get_file_placement_path(self, task_type: str, filename: str) -> str:
        """Get the file placement path for a specific task type."""
        if not self.supports_task_type(task_type):
            raise ValueError(f"Unsupported task type: {task_type}")
        path_template = self.file_placement_mapping[task_type]
        return path_template.format(filename=filename)

    def get_directories_to_create(self, task_type: str) -> List[str]:
        """Get list of directories that need to be created for a task type."""
        if not self.supports_task_type(task_type):
            return []
        return self.directory_creation_mapping.get(task_type, [])

    def create_task_files(self, task_data: Dict[str, Any], task_id: str, temp_dir: str) -> Optional[Dict[str, str]]:
        """Create tabular data reporting specific task files."""
        task_type = task_data.get("task_type")
        if not self.supports_task_type(task_type):
            return None
        if task_type == "simple_data_transfer":
            return self._create_simple_data_transfer_files(task_data, task_id, temp_dir)
        elif task_type == "basic_data_aggregation":
            return self._create_basic_data_aggregation_files(task_data, task_id, temp_dir)
        elif task_type == "simple_calculation_output":
            return self._create_simple_calculation_output_files(task_data, task_id, temp_dir)
        return None

    def supports_task_type(self, task_type: str) -> bool:
        """Check if this provider supports the given task type."""
        return task_type in self.supported_tasks

    def _create_simple_data_transfer_files(self, task_data: Dict[str, Any], task_id: str, temp_dir: str) -> Dict[str, str]:
        """Create files for simple data transfer task."""
        data_content = task_data.get("data_file_content", task_data.get("data_content", "Name,Age,City\nJohn,25,NYC\nJane,30,LA"))
        data_filename = task_data.get("data_filename", f"data_{task_id[:8]}.txt")
        data_file_path = os.path.join(temp_dir, data_filename)
        with open(data_file_path, "w") as f:
            f.write(data_content)
        expected_files = self._create_expected_spreadsheet_file(task_data, task_id, temp_dir, "simple_data_transfer")
        files_info = {"main_file": data_file_path, "main_filename": data_filename}
        if expected_files:
            files_info.update(expected_files)
        return files_info

    def _create_basic_data_aggregation_files(self, task_data: Dict[str, Any], task_id: str, temp_dir: str) -> Dict[str, str]:
        """Create files for basic data aggregation task."""
        numbers_content = task_data.get("data_file_content", task_data.get("numbers_content", "10\n20\n30\n40\n50"))
        data_filename = task_data.get("data_filename", f"numbers_{task_id[:8]}.txt")
        data_file_path = os.path.join(temp_dir, data_filename)
        with open(data_file_path, "w") as f:
            f.write(numbers_content)
        expected_files = self._create_expected_spreadsheet_file(task_data, task_id, temp_dir, "basic_data_aggregation")
        files_info = {"main_file": data_file_path, "main_filename": data_filename}
        if expected_files:
            files_info.update(expected_files)
        return files_info

    def _create_simple_calculation_output_files(self, task_data: Dict[str, Any], task_id: str, temp_dir: str) -> Dict[str, str]:
        """Create files for simple calculation output task."""
        spreadsheet_data = task_data.get("spreadsheet_data", {})
        if not spreadsheet_data:
            spreadsheet_data = {"column_a": [10, 15, 8, 20], "column_b": [25, 30, 15, 40], "row_count": 4}
        domain_context = task_data.get("domain_context", {})
        if domain_context:
            scenario = domain_context.get("scenario", "business_analytics")
            if scenario == "business_analytics":
                headers = ["Revenue", "Costs"]
            elif scenario == "scientific_research":
                headers = ["Sensor_A", "Sensor_B"]
            elif scenario == "education_management":
                headers = ["Test_1", "Test_2"]
            elif scenario == "inventory_logistics":
                headers = ["Stock_A", "Stock_B"]
            else:
                headers = ["Income", "Expenses"]
        else:
            headers = ["Column_A", "Column_B"]
        column_a_values = spreadsheet_data.get("column_a", [10, 15, 8])
        column_b_values = spreadsheet_data.get("column_b", [25, 30, 15])
        data = {headers[0]: column_a_values, headers[1]: column_b_values}
        filename = task_data.get("file_name", f"task_data_{task_id[:8]}.xlsx")
        file_path = os.path.join(temp_dir, filename)
        try:
            import pandas as pd

            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
        except ImportError:
            print("Warning: pandas not available, creating CSV file for LibreOffice Calc")
            csv_filename = filename.replace(".xlsx", ".csv")
            csv_file_path = os.path.join(temp_dir, csv_filename)
            with open(csv_file_path, "w") as f:
                f.write(",".join(headers) + "\n")
                max_rows = max(len(column_a_values), len(column_b_values))
                for i in range(max_rows):
                    row = []
                    row.append(str(column_a_values[i]) if i < len(column_a_values) else "")
                    row.append(str(column_b_values[i]) if i < len(column_b_values) else "")
                    f.write(",".join(row) + "\n")
            file_path = csv_file_path
            filename = csv_filename
        expected_files = self._create_expected_spreadsheet_file(task_data, task_id, temp_dir, "simple_calculation_output")
        files_info = {"main_file": file_path, "main_filename": filename}
        if expected_files:
            files_info.update(expected_files)
        if "expected_total" in task_data:
            expected_filename = f"expected_total_{task_id[:8]}.txt"
            expected_file_path = os.path.join(temp_dir, expected_filename)
            with open(expected_file_path, "w") as f:
                f.write(str(task_data["expected_total"]))
            files_info["gold_standard_file"] = expected_file_path
            files_info["expected_filename"] = expected_filename
        return files_info

    def _create_expected_spreadsheet_file(self, task_data: Dict[str, Any], task_id: str, temp_dir: str, task_type: str) -> Optional[Dict[str, str]]:
        """Create expected spreadsheet file for evaluation."""
        try:
            from openpyxl import Workbook
        except ImportError:
            print("Warning: openpyxl not available, cannot create expected spreadsheet files")
            return None
        evaluation_data = task_data.get("evaluation_data", {})
        wb = Workbook()
        ws = wb.active
        if task_type == "simple_data_transfer":
            target_cell = evaluation_data.get("target_cell", "A1")
            expected_value = evaluation_data.get("expected_value", 0)
            ws[target_cell] = expected_value
        elif task_type == "basic_data_aggregation":
            target_cell = evaluation_data.get("target_cell", "A1")
            expected_value = evaluation_data.get("expected_value", 0)
            ws[target_cell] = expected_value
        elif task_type == "simple_calculation_output":
            target_cell = evaluation_data.get("spreadsheet_target_cell", "A1")
            expected_value = evaluation_data.get("spreadsheet_expected_value", 0)
            ws[target_cell] = expected_value
        expected_filename = "expected_result.xlsx"
        expected_path = os.path.join(temp_dir, expected_filename)
        wb.save(expected_path)
        return {
            "expected_spreadsheet_file": expected_path,
            "expected_spreadsheet_filename": expected_filename,
            "gold_standard_file": expected_path,
            "ground_truth_files": {"expected_spreadsheet_file": expected_path},
        }


class TabularDataReportingConfigProvider(ConfigProviderInterface):
    """Config provider implementation for tabular data reporting tasks."""

    def __init__(self):
        """Initialize tabular data reporting config provider."""
        self.supported_tasks = {
            "simple_data_transfer",
            "basic_data_aggregation",
            "simple_calculation_output",
        }
        self.evaluation_mode_mapping = {
            "simple_data_transfer": "compare_table",
            "basic_data_aggregation": "compare_table",
            "simple_calculation_output": "compare_table",
        }

    def build_setup_steps(
        self,
        task_data: Dict[str, Any],
        s3_urls: Dict[str, str],
        files_created: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """Build setup configuration steps for task initialization."""
        task_type = task_data["task_type"]
        if not self.supports_task_type(task_type):
            return []
        if task_type == "simple_data_transfer":
            return self._build_level1_setup(task_data, s3_urls, files_created)
        elif task_type == "basic_data_aggregation":
            return self._build_level2_setup(task_data, s3_urls, files_created)
        elif task_type == "simple_calculation_output":
            return self._build_level3_setup(task_data, s3_urls, files_created)
        return []

    def get_evaluation_mode(self, task_type: str, level: int) -> str:
        """Get the evaluation mode for a specific task type and level."""
        if not self.supports_task_type(task_type):
            return "compare_text_file"
        return self.evaluation_mode_mapping.get(task_type, "compare_table")

    def supports_task_type(self, task_type: str) -> bool:
        """Check if this provider supports the given task type."""
        return task_type in self.supported_tasks

    def _build_level1_setup(
        self,
        task_data: Dict[str, Any],
        s3_urls: Dict[str, str],
        files_created: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """Build setup steps for Level 1: Simple Data Transfer."""
        steps = []
        steps.append(
            {
                "type": "download",
                "parameters": {
                    "files": [
                        {
                            "url": s3_urls["main_file"],
                            "path": f"/home/user/Desktop/{task_data['data_filename']}",
                        }
                    ]
                },
            }
        )
        steps.append({"type": "launch", "parameters": {"command": ["libreoffice", "--calc"]}})
        steps.append({"type": "sleep", "parameters": {"seconds": 10.0}})
        return steps

    def _build_level2_setup(
        self,
        task_data: Dict[str, Any],
        s3_urls: Dict[str, str],
        files_created: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """Build setup steps for Level 2: Basic Data Aggregation."""
        steps = []
        steps.append(
            {
                "type": "download",
                "parameters": {
                    "files": [
                        {
                            "url": s3_urls["main_file"],
                            "path": f"/home/user/Desktop/{task_data['data_filename']}",
                        }
                    ]
                },
            }
        )
        steps.append({"type": "launch", "parameters": {"command": ["libreoffice", "--calc"]}})
        steps.append({"type": "sleep", "parameters": {"seconds": 10.0}})
        return steps

    def _build_level3_setup(
        self,
        task_data: Dict[str, Any],
        s3_urls: Dict[str, str],
        files_created: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """Build setup steps for Level 3: Simple Calculation and Output."""
        steps = []
        main_filename = files_created.get("main_filename", "task_data.xlsx")
        steps.append(
            {
                "type": "download",
                "parameters": {
                    "files": [
                        {
                            "url": s3_urls["main_file"],
                            "path": f"/home/user/Desktop/{main_filename}",
                        }
                    ]
                },
            }
        )
        steps.append(
            {
                "type": "launch",
                "parameters": {
                    "command": [
                        "libreoffice",
                        "--calc",
                        "--infilter=CSV:44,34,0,1,1",
                        f"/home/user/Desktop/{main_filename}",
                    ]
                },
            }
        )
        steps.append({"type": "sleep", "parameters": {"seconds": 10.0}})
        return steps


class TabularDataReportingEvaluationProvider(EvaluationProviderInterface):
    """Evaluation provider implementation for tabular data reporting tasks."""

    def __init__(self):
        """Initialize tabular data reporting evaluation provider."""
        self.supported_tasks = {
            "simple_data_transfer",
            "basic_data_aggregation",
            "simple_calculation_output",
        }
        self.evaluators = TabularDataReportingEvaluators()

    def build_evaluator_config(
        self,
        task_data: Dict[str, Any],
        files_created: Dict[str, str],
        evaluation_mode: str,
        s3_urls: Dict[str, str],
    ) -> Dict[str, Any]:
        """Build evaluator configuration for the task."""
        task_type = task_data.get("task_type")
        if not self.supports_task_type(task_type):
            return {}
        task_data_with_mode = {**task_data, "evaluation_mode": evaluation_mode}
        if self.evaluators.needs_multi_evaluator(task_type):
            return self.evaluators.build_multi_evaluator_config(task_type, task_data_with_mode, files_created, s3_urls)
        else:
            return self.evaluators.build_single_evaluator_config(task_type, task_data_with_mode, files_created, s3_urls)

    def supports_task_type(self, task_type: str) -> bool:
        """Check if this provider supports the given task type."""
        return task_type in self.supported_tasks

    def get_evaluator_instance(self):
        """Get the evaluator instance for this category."""
        return self.evaluators


__all__ = [
    "TabularDataReportingFileProvider",
    "TabularDataReportingConfigProvider",
    "TabularDataReportingEvaluationProvider",
]
